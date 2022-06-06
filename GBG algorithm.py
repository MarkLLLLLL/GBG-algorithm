# -*- coding: utf-8 -*-
# The greedy-based generation algorithm is coded with python language
from abaqus import *
from abaqusConstants import *
import regionToolset
import os
import math
import time
import os
from xlwt import Workbook
import random
import numpy as np
import operator
import os.path
start = time.clock()
import xlrd
from openpyxl import load_workbook
import displayGroupMdbToolset as dgm

def interCheck(point,center):
    sign = True
    for p in center:
        if sqrt((point[0] - p[0]) ** 2 + (point[1] - p[1]) ** 2) <= (lmin+point[2]+p[2]):
            sign = False
            break
    return sign

def volumefraction(length2, width2, n, radius3, vf2):
    sign2 = True
    s=0
    for i in range(n):
        s=s+math.pi*(radius3[i])**2
    if s/(length2*width2)>vf2 and abs(s/(length2*width2)-vf2)>0.5*math.pi*radius*radius :
        sign2 = False
    return sign2


for elastic_caluation_number in range (1,2):
    Mdb()
    a = mdb.models['Model-1'].rootAssembly
    session.viewports['Viewport: 1'].setValues(displayedObject=a)
    print elastic_caluation_number
    path = "C:\\Users\\lmz\\Desktop"
    w=Workbook()
    ws=w.add_sheet('Fiber coordinate')
    myModel = mdb.models["Model-1"]
    width = 0.13
    length = 0.13
    mySketch = mdb.models['Model-1'].ConstrainedSketch(name='sketch_1', sheetSize=200.0)
    mySketch.rectangle(point1=(0,0), point2=(length, width))
    myPart = myModel.Part(name='Part-1', dimensionality=TWO_D_PLANAR, type=DEFORMABLE_BODY)
    myPart.BaseShell(sketch=mySketch)
    del myModel.sketches['sketch_1']
    for rve in range(1,2):
        lmin = 0.000182
        lmax = 0.0008
        vf = 0.7
        count = 0
        countmax= 1
        n = 1
        base =0 
        m = 23 
        radius = 0.0026
        p0=0.99
        nmax=round(width*length*vf/math.pi/radius/radius*1.1) 
        print nmax
        radiuslis = np.random.normal(loc=radius, scale=0.0001553, size=nmax)
        mySketch = myModel.ConstrainedSketch(name='sketch_2', sheetSize=200)
        xbase = random.uniform(9*length/20, 11*length/20)
        ybase = random.uniform(9*width/20, 11*width/20)
        center = [[xbase, ybase, radiuslis[0]]]
        while True:
            pjum=random.random()
            xy=[]
            array1=[]
            for k in range(m):
                l = random.uniform(lmin+radiuslis[base]+radiuslis[n], lmax+radiuslis[base]+radiuslis[n])
                # murandom=random.randint(0,1)
                # mu=murandom*math.pow((lmax-l+radiuslis[base]+radiuslis[n])/(lmax-lmin),0.9)\
                # *math.pi*0.5+(1-murandom)*(math.pi-math.pow((lmax-l+radiuslis[base]+radiuslis[n])/(lmax-lmin),0.9)*math.pi*0.5)
                # deta=random.normalvariate(mu,DETASIGMA)
                # fuhaorandom=random.random()
                # if fuhaorandom<0.5:
                #     deta=deta
                # else:
                #     deta=math.pi+deta
                
                # print deta
                # fuhaorandom=random.random()
                # if fuhaorandom<0.5:
                #     deta=deta
                # else:
                #     deta=math.pi+deta
                # if deta>4*math.pi:
                #     deta=deta-4*math.pi
                # elif deta<-2*math.pi:
                #     deta=deta+4*math.pi
                
                # if deta>2*math.pi:
                #     deta=deta-2*math.pi
                # elif deta<0:
                #     deta=deta+2*math.pi
                
                deta = random.uniform(0,2*math.pi)
                tx = xbase + l* math.cos(deta)
                ty = ybase + l* math.sin(deta)
                xy.append([tx,ty,radiuslis[n]])
            if pjum<=p0:
                for i in xy:
                    for j in center:
                        if sqrt((i[0]-j[0])**2+(i[1]-j[1])**2)==l:
                            sij=0
                            array1.append(sij)
                        elif sqrt((i[0]-j[0])**2+(i[1]-j[1])**2)<(lmin+i[2]+j[2]):
                            sij=float('inf')
                            array1.append(sij)
                        elif lmin+i[2]+j[2]<sqrt((i[0]-j[0])**2+(i[1]-j[1])**2)<l:
                            sij=(l-sqrt((i[0]-j[0])**2+(i[1]-j[1])**2))/(l-i[2]-j[2])
                            array1.append(sij)
                        elif sqrt((i[0]-j[0])**2+(i[1]-j[1])**2)>l:
                            sij=(sqrt((i[0]-j[0])**2+(i[1]-j[1])**2)-l)/(l-i[2]-j[2])
                            array1.append(sij)            
                        
                    
                
            else:
                for i in xy:
                    for j in center:
                        if sqrt((i[0]-j[0])**2+(i[1]-j[1])**2)>lmin+i[2]+j[2]:
                            sij=0
                            array1.append(sij)
                        else:
                            sij=float('inf')
                            array1.append(sij)                    
            sfit=[]
            for i in range(m):
                summm=0
                if (len(center)-1)==0:
                    summm=summm+array1[i*len(center)]
                else:
                    for j in range(i*len(center),(i+1)*len(center)):
                        summm=summm+array1[j]
                sfit.append(summm)
            min_index, min_number = min(enumerate(sfit), key=operator.itemgetter(1))
            txy=xy[min_index]
            x=txy[0]
            y=txy[1]
            radius02=txy[2]
            flag1=0
            flag2=0
            flag3=0
            sign1=True
            sign2=True
            sign3=True
            if(x<radius02)and((y-radius02)*(width-radius02-y)>0):
                flag1=1
                x1=x+length
                y1=y
            elif(x>length-radius02)and((y-radius02)*(width-radius02-y)>0):
                flag1=1
                x1=x-length
                y1=y
            elif(y< radius02)and((x- radius02)*(length- radius02-x)>0):
                flag1=1
                x1=x
                y1=y+width
            elif(y>width- radius02)and((x- radius02)*(length- radius02-x)>0):
                flag1=1
                x1=x
                y1=y-width          
            elif(x< radius02)and(y< radius02):
                flag1=1     
                x1=x+length
                y1=y
                flag2=1     
                x2=x
                y2=y+width      
                flag3=1
                x3=x+length
                y3=y+width      
            elif(x< radius02)and(y>(width- radius02)):
                flag1=1     
                x1=x+length
                y1=y
                flag2=1     
                x2=x
                y2=y-width
                flag3=1         
                x3=x+length
                y3=y-width          
            elif(x>(length- radius02))and(y< radius02):       
                flag1=1
                x1=x-length
                y1=y
                flag2=1     
                x2=x
                y2=y+width  
                flag3=1     
                x3=x-length
                y3=y+width          
            elif(x>(length- radius02))and(y>(width- radius02)):               
                flag1=1 
                x1=x
                y1=y-width          
                flag2=1 
                x2=x-length
                y2=y    
                flag3=1     
                x3=x-length
                y3=y-width
            sign = interCheck([x,y,radius02],center)
            if flag1==1:
                sign1 = interCheck([x1,y1,radius02],center)
            if flag2==1:
                sign2 = interCheck([x2,y2,radius02],center)
            if flag3==1:
                sign3 = interCheck([x3,y3,radius02],center)
            
            if sign and sign1 and sign2 and sign3 and (0-radius02<x<width+radius02) and (0-radius02<y<length+radius02):
                center.append([x,y,radius02])
                if flag1==1:
                    center.append((x1,y1,radius02))
                if flag2==1:
                    center.append((x2,y2,radius02))  
                if flag3==1:
                    center.append((x3,y3,radius02))  
                n += 1
                if n % 30 == 0:
                    print n
            else:
                count += 1
                if count>=countmax:
                    base += 1
                    if base<len(center):
                        xbase = center[base][0]
                        ybase = center[base][1]
                        count = 0
                    else:
                        break
        for i in range (500):
            signvtrue = volumefraction(length, width, n-i-1, radiuslis, 0.600)
            if (signvtrue):   
                break
            elif (not signvtrue):
                removei=random.randint(0,len(center)-1)
                del_index=[]
                radius02=center[removei][2]
                del center[removei]
                radiuslis = np.delete(radiuslis, removei)
                for iiiii in range(len(center)):
                    if radius02==center[iiiii][2]:
                        del_index.append(iiiii)
                center=[iii for num,iii in enumerate(center) if num not in del_index]
                
            
        
        
        count2=0
        for each in center:
            ws.write(count2,rve*2-2, each[0])
            ws.write(count2,rve*2-1, each[1])
            count2=count2+1
        localPath=os.path.join(path, 'Fiber coordinate.xls')
        w.save(localPath)
        for p in center:
            x,y,z = p[0], p[1], p[2]
            mySketch.CircleByCenterPerimeter(center=(x, y), point1=(x+z, y))
        myPart.PartitionFaceBySketch(faces=myPart.faces[0:1], sketch=mySketch)
        del myModel.sketches['sketch_2']
        print n-i
        print len(center)
        print base
        s1=0
        s2=0
        for i1 in range(n):
            s1=s1+math.pi*(radiuslis[i1])**2
        print s1/(length*width)
        for i2 in range(n-i):
            s2=s2+math.pi*(center[i2][2])**2
        print s2/(length*width)
        print rve
        end=time.clock()
        center=[]
        print('running time:%s seconds'%(end-start))
    
    modelName= 'Model-1'
    instanceName= 'Part-1-1'
    mdb.models['Model-1'].Material(name='Matrix')
    mdb.models['Model-1'].materials['Matrix'].Elastic(table=((3350.0, 0.35), ))
    mdb.models['Model-1'].Material(name='Fiber')
    mdb.models['Model-1'].materials['Fiber'].Elastic(table=((74000.0, 0.2), ))
    mdb.models['Model-1'].HomogeneousSolidSection(name='Matrix', material='Matrix', 
        thickness=None)
    mdb.models['Model-1'].HomogeneousSolidSection(name='Fiber', material='Fiber', 
        thickness=None)
    a = mdb.models[modelName].rootAssembly
    for i in a.sets.keys():
        if i==instanceName:
            continue
        else:
            del a.sets[i]
        
    
    a = mdb.models[modelName]
    for i in a.constraints.keys():
        if i==instanceName:
            continue
        else:
            del a.constraints[i]
        
    p = mdb.models['Model-1'].parts['Part-1']
    f = p.faces
    for i in range(500):
        x=random.uniform(0,length)
        y=random.uniform(0,length)
        face=f.findAt((x,y,0),)
        faces = f.findAt(((x,y,0),),)
        if face.getSize()>math.pi*0.0038*0.0038:
            p.Set(faces=faces, name='Set-Matrix')
            break
        
    
    region = regionToolset.Region(faces=faces)
    p = mdb.models['Model-1'].parts['Part-1']
    p.SectionAssignment(region=region, sectionName='Matrix', offset=0.0, 
        offsetType=MIDDLE_SURFACE, offsetField='', 
        thicknessAssignment=FROM_SECTION)
    data = xlrd.open_workbook('C:\\Users\\lmz\\Desktop\\Fiber coordinate.xls')
    sh=data.sheet_by_name('Fiber coordinate')
    for p in range(sh.nrows):
        x,y,z = sh.cell(p, 0).value,sh.cell(p, 1).value,0
        if x<0:
            x=0
        elif x>length:
            x=length
        
        if y<0:
            y=0
        elif y>length:
            y=length
        
        a=((x,y,z),)
        facesa=f.findAt(a)
        if p==0:
            b=facesa
        else:
            b+=facesa
        
    
    region = regionToolset.Region(faces=b)
    p = mdb.models['Model-1'].parts['Part-1']
    p.SectionAssignment(region=region, sectionName='Fiber', offset=0.0, 
        offsetType=MIDDLE_SURFACE, offsetField='', 
        thicknessAssignment=FROM_SECTION)
    a1 = mdb.models['Model-1'].rootAssembly
    a1.DatumCsysByDefault(CARTESIAN)
    p = mdb.models['Model-1'].parts['Part-1']
    a1.Instance(name='Part-1-1', part=p, dependent=ON)
    mdb.models['Model-1'].StaticStep(name='Step-1', previous='Initial')
    session.viewports['Viewport: 1'].assemblyDisplay.setValues(step='Step-1')
    p = mdb.models['Model-1'].parts['Part-1']
    session.viewports['Viewport: 1'].setValues(displayedObject=p)
    session.viewports['Viewport: 1'].partDisplay.setValues(sectionAssignments=OFF, 
        engineeringFeatures=OFF, mesh=ON)
    session.viewports['Viewport: 1'].partDisplay.meshOptions.setValues(
        meshTechnique=ON)
    p = mdb.models['Model-1'].parts['Part-1']
    p.seedPart(size=0.00065, deviationFactor=0.1, minSizeFactor=0.1)
    p = mdb.models['Model-1'].parts['Part-1']
    p.generateMesh()
    import mesh
    elemType1 = mesh.ElemType(elemCode=CPE4, elemLibrary=STANDARD)
    elemType2 = mesh.ElemType(elemCode=CPE3, elemLibrary=STANDARD, 
        secondOrderAccuracy=OFF, distortionControl=DEFAULT)
    p = mdb.models['Model-1'].parts['Part-1']
    f = p.faces
    faces = f.getByBoundingBox(-0.01,-0.01,-0.01,0.2,0.2,0.2)
    pickedRegions =(faces, )
    p.setElementType(regions=pickedRegions, elemTypes=(elemType1, elemType2))
    modelName='Model-1'
    instanceName= 'Part-1-1'
    node = mdb.models[modelName].rootAssembly.instances[instanceName].nodes
    Xmin = 9999
    Xmax = -9999
    Ymin = 9999
    Ymax = -9999
    c1=[]
    c2=[]
    c3=[]
    c4=[]
    coc1={}
    coc2={}
    coc3={}
    coc4={}
    a = mdb.models[modelName].rootAssembly
    for i in a.features.keys():
        if i==instanceName:
            continue
        else:
            del a.features[i]
    xx=math.cos(22.5/360*2*math.pi)
    yy=math.sin(22.5/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-1', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(1.0, 1.0, 0.0))
    datum1 = mdb.models['Model-1'].rootAssembly.datums[4]
    xx=math.cos(45.0/360*2*math.pi)
    yy=math.sin(45.0/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-2', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(1.0, 2.0, 0.0))
    datum2 = mdb.models['Model-1'].rootAssembly.datums[5]
    xx=math.cos(67.5/360*2*math.pi)
    yy=math.sin(67.5/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-3', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(-1.0, 1.0, 0.0))
    datum3 = mdb.models['Model-1'].rootAssembly.datums[6]
    xx=math.cos(90.0/360*2*math.pi)
    yy=math.sin(90.0/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-4', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(-1.0, 1.0, 0.0))
    datum4 = mdb.models['Model-1'].rootAssembly.datums[7]
    xx=math.cos(112.5/360*2*math.pi)
    yy=math.sin(112.5/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-5', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(-10.0, 1.0, 0.0))
    datum5 = mdb.models['Model-1'].rootAssembly.datums[8]
    xx=math.cos(135.0/360*2*math.pi)
    yy=math.sin(135.0/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-6', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(-10.0, 1.0, 0.0))
    datum6 = mdb.models['Model-1'].rootAssembly.datums[9]
    xx=math.cos(157.5/360*2*math.pi)
    yy=math.sin(157.5/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-7', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(-10.0, 1.0, 0.0))
    datum7 = mdb.models['Model-1'].rootAssembly.datums[10]
    xx=math.cos(180.0/360*2*math.pi)
    yy=math.sin(180.0/360*2*math.pi)
    a = mdb.models['Model-1'].rootAssembly
    a.DatumCsysByThreePoints(name='Datum csys-8', coordSysType=CARTESIAN, origin=(
        0.0, 0.0, 0.0), point1=(xx, yy, 0.0), point2=(-1.0, -1.0, 0.0))
    datum8 = mdb.models['Model-1'].rootAssembly.datums[11]
    for i in range(len(node)):
        x = node[i].coordinates[0]
        y = node[i].coordinates[1]
        if Xmin > x:
            Xmin = x
        elif Xmax < x:
            Xmax = x
        if Ymin > y:
            Ymin = y
        elif Ymax < y:
            Ymax = y
    eps1 = abs(Xmax - Xmin) * 0.00001
    eps2 = abs(Xmax - Xmin) * 0.0001
    BX = Xmax - Xmin
    BY = Ymax - Ymin
    Dispx=BX*0.1
    Dispy=BY*0.1
    upperName= instanceName.upper()
    node = mdb.models[modelName].rootAssembly.instances[instanceName].nodes
    node_v1 = node[1:1]
    node_v2 = node[1:1]
    node_v3 = node[1:1]
    node_v4 = node[1:1]
    node_E5 = node[1:1]
    node_E6 = node[1:1]
    node_E9 = node[1:1]
    node_E10 = node[1:1]
    for i in range(len(node)):
        x = node[i].coordinates[0]
        y = node[i].coordinates[1]
        z = node[i].coordinates[2]
        if abs(x - Xmin) < eps1 and abs(y - Ymin) < eps1:
            node_v1 = node[i:i + 1]
            c1.insert(0,node[i].label)
            coc1[node[i].label]=[node[i].coordinates[0],node[i].coordinates[1],node[i].coordinates[2]]
            continue
        if abs(x - Xmax) < eps1 and abs(y - Ymin) < eps1:
            node_v2 = node[i:i + 1]
            c2.insert(0,node[i].label)
            coc2[node[i].label]=[node[i].coordinates[0],node[i].coordinates[1],node[i].coordinates[2]]
            continue
        if abs(x - Xmax) < eps1 and abs(y - Ymax) < eps1:
            node_v3 = node[i:i + 1]
            c3.insert(0,node[i].label)
            coc3[node[i].label]=[node[i].coordinates[0],node[i].coordinates[1],node[i].coordinates[2]]
            continue
        if abs(x - Xmin) < eps1 and abs(y - Ymax) < eps1:
            node_v4 = node[i:i + 1]
            c4.insert(0,node[i].label)
            coc4[node[i].label]=[node[i].coordinates[0],node[i].coordinates[1],node[i].coordinates[2]]
            continue
        if abs(x - Xmin) < eps1 and abs(y - Ymin) > eps2 and abs(y - Ymax) > eps2:
            node_E5 = node_E5 + node[i:i + 1]
            continue
        if abs(x - Xmax) < eps1 and abs(y - Ymin) > eps2 and abs(y - Ymax) > eps2:
            node_E6 = node_E6 + node[i:i + 1]
            continue
        if abs(y - Ymin) < eps1 and abs(x - Xmin) > eps2 and abs(x - Xmax) > eps2:
            node_E9 = node_E9 + node[i:i + 1]
            continue
        if abs(y - Ymax) < eps1 and abs(x - Xmin) > eps2 and abs(x - Xmax) > eps2:
            node_E10 = node_E10 + node[i:i + 1]
            continue
    mdb.models[modelName].rootAssembly.Set(name = 'Master Node 1', nodes = node_v1)
    mdb.models[modelName].rootAssembly.Set(name = 'Master Node 2', nodes = node_v2)
    mdb.models[modelName].rootAssembly.Set(name = 'Master Node 3', nodes = node_v3)
    mdb.models[modelName].rootAssembly.Set(name = 'Master Node 4', nodes = node_v4)
    mdb.models[modelName].rootAssembly.Set(name = 'edge5 V', nodes = node_E5)
    mdb.models[modelName].rootAssembly.Set(name = 'edge6 VI', nodes = node_E6)
    mdb.models[modelName].rootAssembly.Set(name = 'edge9 IX', nodes = node_E9)
    mdb.models[modelName].rootAssembly.Set(name = 'edge10 X', nodes = node_E10)
    print len(node), 'node set:'
    print 'E5', len(node_E5), len(node_E6), len(node_E9), len(node_E10)
    Ele_E5 = []
    Ele_E6 = []
    Ele_E9 = []
    Ele_E10 = []
    element = mdb.models[modelName].rootAssembly.instances[instanceName].elements
    for i in range(len(element)):
        E_E5 = [i]
        E_E6 = [i]
        E_E9 = [i]
        E_E10 = [i]
        for j in element[i].connectivity:
            x = node[j].coordinates[0]
            y = node[j].coordinates[1]
            if abs(x - Xmin) < eps1:
                E_E5.append(j)
            elif abs(x - Xmax) < eps1:
                E_E6.append(j)
            if abs(y - Ymin) < eps1:
                E_E9.append(j)
            elif abs(y - Ymax) < eps1:
                E_E10.append(j)
        if len(E_E5) == 3:
            Ele_E5.append(E_E5)
        elif len(E_E6) == 3:
            Ele_E6.append(E_E6)
        elif len(E_E9) == 3:
            Ele_E9.append(E_E9)
        elif len(E_E10) == 3:
            Ele_E10.append(E_E10)
    Ele_List = [
        Ele_E5,
        Ele_E6,
        Ele_E9,
        Ele_E10]
    for i in range(len(Ele_E5)):
        mdb.models[modelName].rootAssembly.Set(name = 'Ele_E5_' + str(i), elements = element[Ele_E5[i][0]:Ele_E5[i][0] + 1])
    
    nodeset = mdb.models[modelName].rootAssembly.sets['edge5 V'].nodes
    nodeset = nodeset + mdb.models[modelName].rootAssembly.sets['edge6 VI'].nodes
    nodeset = nodeset + mdb.models[modelName].rootAssembly.sets['edge9 IX'].nodes
    nodeset = nodeset + mdb.models[modelName].rootAssembly.sets['edge10 X'].nodes
    nodeset = nodeset + mdb.models[modelName].rootAssembly.sets['Master Node 1'].nodes
    nodeset = nodeset + mdb.models[modelName].rootAssembly.sets['Master Node 2'].nodes
    nodeset = nodeset + mdb.models[modelName].rootAssembly.sets['Master Node 3'].nodes
    nodeset = nodeset + mdb.models[modelName].rootAssembly.sets['Master Node 4'].nodes
    
    print 'Number of node set for interpolation:', len(nodeset)
    for i in range(len(nodeset)):
        mdb.models[modelName].rootAssembly.Set(name = 'Node' + str(nodeset[i].label - 1), nodes = nodeset[i:i + 1])
    
    print 'Creat Mesh sets Complete'
    
    mdb.models[modelName].rootAssembly.ReferencePoint(point = (Xmax+0.1, 0.0, 0))
    mdb.models[modelName].rootAssembly.ReferencePoint(point = (0.0, Ymax+0.1, 0))
    mdb.models[modelName].rootAssembly.ReferencePoint(point = (Xmax+0.2, 0.0, 0))
    mdb.models[modelName].rootAssembly.ReferencePoint(point = (0, Ymax+0.2, 0))
    mdb.models[modelName].rootAssembly.ReferencePoint(point = (Xmax+0.3, 0, 0))
    mdb.models[modelName].rootAssembly.ReferencePoint(point = (0, Ymax+0.3, 0))
    mdb.models[modelName].rootAssembly.features.changeKey(fromName = 'RP-1', toName = 'Fx')
    mdb.models[modelName].rootAssembly.features.changeKey(fromName = 'RP-2', toName = 'Fy')
    mdb.models[modelName].rootAssembly.features.changeKey(fromName = 'RP-3', toName = 'Shear_xy')
    mdb.models[modelName].rootAssembly.features.changeKey(fromName = 'RP-4', toName = 'Shear_xyB')
    mdb.models[modelName].rootAssembly.features.changeKey(fromName = 'RP-5', toName = 'Yx')
    mdb.models[modelName].rootAssembly.features.changeKey(fromName = 'RP-6', toName = 'Yy')
    keyRefPointList = mdb.models[modelName].rootAssembly.referencePoints.keys()
    refPoint = (mdb.models[modelName].rootAssembly.referencePoints[keyRefPointList[5]],)
    mdb.models[modelName].rootAssembly.Set(referencePoints = refPoint, name = 'Constraints Driver Fx')
    refPoint = (mdb.models[modelName].rootAssembly.referencePoints[keyRefPointList[4]],)
    mdb.models[modelName].rootAssembly.Set(referencePoints = refPoint, name = 'Constraints Driver Fy')
    refPoint = (mdb.models[modelName].rootAssembly.referencePoints[keyRefPointList[3]],)
    mdb.models[modelName].rootAssembly.Set(referencePoints = refPoint, name = 'Constraints Driver Shear_xy')
    refPoint = (mdb.models[modelName].rootAssembly.referencePoints[keyRefPointList[2]],)
    mdb.models[modelName].rootAssembly.Set(referencePoints = refPoint, name = 'Constraints Driver Shear_xyB')
    refPoint = (mdb.models[modelName].rootAssembly.referencePoints[keyRefPointList[1]],)
    mdb.models[modelName].rootAssembly.Set(referencePoints = refPoint, name = 'Constraints Driver Yx')
    refPoint = (mdb.models[modelName].rootAssembly.referencePoints[keyRefPointList[0]],)
    mdb.models[modelName].rootAssembly.Set(referencePoints = refPoint, name = 'Constraints Driver Yy')
    BX = Xmax - Xmin
    BY = Ymax - Ymin
    node = mdb.models[modelName].rootAssembly.instances[instanceName].nodes
    Ele_E5 = Ele_List[0]
    Ele_E6 = Ele_List[1]
    Ele_E9 = Ele_List[2]
    Ele_E10 = Ele_List[3]
    mdb.models[modelName].Equation(name = 'U2-U1 Master node on u', terms = ((1, 'Master Node 2', 1), (-1, 'Master Node 1', 1), (-1, 'Constraints Driver Fx', 1), (1, 'Constraints Driver Shear_xy', 1)))
    mdb.models[modelName].Equation(name = 'U2-U1 Master node on v', terms = ((1, 'Master Node 2', 2), (-1, 'Master Node 1', 2), (-1, 'Constraints Driver Yx', 2)))
    mdb.models[modelName].Equation(name = 'U4-U1 Master node on u', terms = ((1, 'Master Node 4', 1), (-1, 'Master Node 1', 1), (-1, 'Constraints Driver Fy', 1), (-1, 'Constraints Driver Shear_xyB', 1)))
    mdb.models[modelName].Equation(name = 'U4-U1 Master node on v', terms = ((1, 'Master Node 4', 2), (-1, 'Master Node 1', 2), (-1, 'Constraints Driver Yy', 2)))
    maxIndex = mdb.models[modelName].rootAssembly.sets['edge6 VI'].nodes
    for i in range(len(maxIndex)):
        x = maxIndex[i].coordinates[0]
        y = maxIndex[i].coordinates[1]
        z = maxIndex[i].coordinates[2]
        Mx = Xmin
        My = y
        Mz = 0
        for j in range(len(Ele_E5)):
            xi = node[Ele_E5[j][1]].coordinates[0]
            yi = node[Ele_E5[j][1]].coordinates[1]
            zi = node[Ele_E5[j][1]].coordinates[2]
            xj = node[Ele_E5[j][2]].coordinates[0]
            yj = node[Ele_E5[j][2]].coordinates[1]
            zj = node[Ele_E5[j][2]].coordinates[2]
            Lij = sqrt((xj - xi) * (xj - xi) + (yj - yi) * (yj - yi) + (zj - zi) * (zj - zi))
            Ljx = sqrt((xj - Mx) * (xj - Mx) + (yj - My) * (yj - My) + (zj - Mz) * (zj - Mz))
            Lix = sqrt((xi - Mx) * (xi - Mx) + (yi - My) * (yi - My) + (zi - Mz) * (zi - Mz))
            if abs(Lij - Ljx - Lix) < 1e-05:
                node_i = Ele_E5[j][1]
                node_j = Ele_E5[j][2]
                break
                continue
        
        setname = 'edge6 VI ' + str(i)
        mdb.models[modelName].rootAssembly.Set(name = setname, nodes = maxIndex[i:i + 1])
        eqName = 'edge6_' + str(i) + ' on u'
        mdb.models[modelName].Equation(name = eqName, terms = ((1, setname, 1), (-Ljx / Lij, 'Node' + str(node_i), 1), (-Lix / Lij, 'Node' + str(node_j), 1), (-1, 'Constraints Driver Fx', 1), (1, 'Constraints Driver Shear_xy', 1)))
        eqName = 'edge6_' + str(i) + ' on v'
        mdb.models[modelName].Equation(name = eqName, terms = ((1, setname, 2), (-Ljx / Lij, 'Node' + str(node_i), 2), (-Lix / Lij, 'Node' + str(node_j), 2), (-1, 'Constraints Driver Yx', 2)))
    maxIndex = mdb.models[modelName].rootAssembly.sets['edge10 X'].nodes
    for i in range(len(maxIndex)):
        x = maxIndex[i].coordinates[0]
        y = maxIndex[i].coordinates[1]
        z = maxIndex[i].coordinates[2]
        Mx = x
        My = Ymin
        Mz = 0
        for j in range(len(Ele_E9)):
            xi = node[Ele_E9[j][1]].coordinates[0]
            yi = node[Ele_E9[j][1]].coordinates[1]
            zi = node[Ele_E9[j][1]].coordinates[2]
            xj = node[Ele_E9[j][2]].coordinates[0]
            yj = node[Ele_E9[j][2]].coordinates[1]
            zj = node[Ele_E9[j][2]].coordinates[2]
            Lij = sqrt((xj - xi) * (xj - xi) + (yj - yi) * (yj - yi) + (zj - zi) * (zj - zi))
            Ljx = sqrt((xj - Mx) * (xj - Mx) + (yj - My) * (yj - My) + (zj - Mz) * (zj - Mz))
            Lix = sqrt((xi - Mx) * (xi - Mx) + (yi - My) * (yi - My) + (zi - Mz) * (zi - Mz))
            if abs(Lij - Ljx - Lix) < 1e-05:
                node_i = Ele_E9[j][1]
                node_j = Ele_E9[j][2]
                break
                continue
        
        setname = 'edge10 X ' + str(i)
        mdb.models[modelName].rootAssembly.Set(name = setname, nodes = maxIndex[i:i + 1])
        eqName = 'edge10_' + str(i) + ' on u'
        mdb.models[modelName].Equation(name = eqName, terms = ((1, setname, 1), (-Ljx / Lij, 'Node' + str(node_i), 1), (-Lix / Lij, 'Node' + str(node_j), 1), (-1, 'Constraints Driver Fy', 1), (-1, 'Constraints Driver Shear_xyB', 1)))
        eqName = 'edge10_' + str(i) + ' on v'
        mdb.models[modelName].Equation(name = eqName, terms = ((1, setname, 2), (-Ljx / Lij, 'Node' + str(node_i), 2), (-Lix / Lij, 'Node' + str(node_j), 2), (-1, 'Constraints Driver Yy', 2)))
